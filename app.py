import streamlit as st
import pdfplumber
import re
import openpyxl
from io import BytesIO

# --- PAGE CONFIG ---
st.set_page_config(page_title="DBR Auto-Tool", layout="centered")

# --- 1. MAPPING LOGIC ---
def get_product_code(text):
    t = text.lower() if text else ""
    if "card" in t: return 8
    if "auto" in t or "car" in t or "lease" in t or "ijara" in t: return 2
    if "personal" in t or "finance" in t or "micro" in t: return 6
    if "running" in t or "od" in t or "cash line" in t: return 34
    return text

def get_term_code(text):
    t = text.lower() if text else ""
    if "card" in t or "running" in t or "cash line" in t: return "E"
    return "T"

# --- 2. PDF PARSER ---
def parse_tasdeeq_pdf(file_obj):
    data = {
        "Name": "", "CNIC": "", "DOB": "", "Loans": []
    }
    try:
        with pdfplumber.open(file_obj) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"

            # --- Personal Info ---
            name_match = re.search(r"Name:\s*(.*?)(?=\s+(Father|Gender|CNIC)|$)", full_text, re.IGNORECASE)
            if name_match: data["Name"] = name_match.group(1).strip()

            cnic_match = re.search(r"\d{5}-\d{7}-\d", full_text)
            if cnic_match: data["CNIC"] = cnic_match.group(0)

            dob_match = re.search(r"Date of Birth:\s*(\d{2}/\d{2}/\d{4})", full_text)
            if dob_match: data["DOB"] = dob_match.group(1)

            # --- Loan Parsing ---
            lines = full_text.split('\n')
            current_loan = {}
            capture_overdues = False
            
            # Patterns
            new_loan_pattern = re.compile(r"^\d+\s?-\s?[A-Za-z]") 
            garbage_pattern = re.compile(r"^\d+\s?-\s?\d+$")

            for line in lines:
                line = line.strip()
                
                # Detect New Loan Header
                if new_loan_pattern.match(line) and not garbage_pattern.match(line):
                    if current_loan: data["Loans"].append(current_loan)
                    bank_part = line.split("-", 1)[1].strip() if "-" in line else line
                    current_loan = {
                        "Bank": bank_part, "Limit": 0, "Outstanding": 0, "MinDue": 0,
                        "30": 0, "60": 0, "90": 0, "Start": "", "End": ""
                    }
                    capture_overdues = False

                if current_loan:
                    if "Loan Limit:" in line:
                        val = re.search(r"[\d,]+", line.split("Limit:")[-1])
                        if val: current_loan["Limit"] = int(val.group(0).replace(",", ""))
                    if "Outstanding Balance:" in line:
                        val = re.search(r"[\d,]+", line.split("Balance:")[-1])
                        if val: current_loan["Outstanding"] = int(val.group(0).replace(",", ""))
                    if "Min Amount Due:" in line:
                        val = re.search(r"[\d,]+", line.split("Due:")[-1])
                        if val: current_loan["MinDue"] = int(val.group(0).replace(",", ""))
                    if "Facility Date:" in line:
                        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", line.split("Facility Date:")[-1])
                        if date_match: current_loan["Start"] = date_match.group(1)
                    if "Maturity Date:" in line:
                        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", line.split("Maturity Date:")[-1])
                        if date_match: current_loan["End"] = date_match.group(1)
                    
                    # Overdue Table Logic
                    if "SUMMARY OF OVERDUES" in line: capture_overdues = True
                    if capture_overdues and line.startswith("Times"):
                        nums = re.findall(r"\d+", line)
                        # Expecting 30+, 60+, 90+
                        if len(nums) >= 3:
                            current_loan["30"] = int(nums[0])
                            current_loan["60"] = int(nums[1])
                            current_loan["90"] = int(nums[2])
                        capture_overdues = False 

            if current_loan: data["Loans"].append(current_loan)

    except Exception as e:
        st.error(f"Error Parsing PDF: {e}")
    
    return data

# --- 3. EXCEL FILLER ---
def fill_excel_template(customer_data, template_path="Template.xlsx"):
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        st.error("⚠️ CRITICAL ERROR: 'Template.xlsx' was not found in the GitHub repo.")
        return None

    ws = wb["INPUT"] # Target the INPUT sheet

    # Fill Identity
    ws['C6'] = customer_data.get("Name", "")
    ws['C7'] = customer_data.get("CNIC", "")
    ws['C8'] = customer_data.get("DOB", "")
    
    # Defaults
    ws['H7'] = 0
    ws['H8'] = 0
    ws['C12'] = "[SELECT]"
    ws['C13'] = "Salaried"

    # Fill Grid (Start Row 19)
    start_row = 19
    current_row = start_row

    for loan in customer_data["Loans"]:
        # Logic to map Bank Name -> Product Code
        p_code = get_product_code(loan["Bank"])
        term_code = get_term_code(loan["Bank"])

        # Write to Cells (Column Indices: A=1, B=2, C=3...)
        ws.cell(row=current_row, column=2).value = "N"          # Col B
        ws.cell(row=current_row, column=3).value = p_code       # Col C
        ws.cell(row=current_row, column=4).value = term_code    # Col D
        ws.cell(row=current_row, column=5).value = loan["Limit"] # Col E
        # F is empty
        ws.cell(row=current_row, column=7).value = loan["Outstanding"] # Col G
        ws.cell(row=current_row, column=8).value = loan["MinDue"]      # Col H
        
        ws.cell(row=current_row, column=9).value = loan["30"]   # Col I
        ws.cell(row=current_row, column=10).value = loan["60"]  # Col J
        ws.cell(row=current_row, column=11).value = loan["90"]  # Col K

        ws.cell(row=current_row, column=12).value = loan["Start"] # Col L
        ws.cell(row=current_row, column=13).value = loan["End"]   # Col M

        current_row += 1

    # Save to Buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- 4. UI LAYOUT ---
st.title("🏦 TASDEEQ -> Excel Tool")
st.markdown("Upload the PDF. Get the filled Excel file.")

uploaded_file = st.file_uploader("Drop Tasdeeq PDF Here", type="pdf")

if uploaded_file:
    with st.spinner("Reading PDF..."):
        # 1. Parse
        data = parse_tasdeeq_pdf(uploaded_file)
        
        if data["Loans"]:
            st.success(f"✅ Extracted {len(data['Loans'])} loans for **{data['Name']}**")
            
            # 2. Fill Excel
            excel_file = fill_excel_template(data)
            
            if excel_file:
                # 3. Create Safe Filename
                safe_name = str(data['Name']).replace("/", "_").replace(" ", "_")
                file_name = f"DBR_{safe_name}.xlsx"

                # 4. Download Button
                st.download_button(
                    label="⬇️ DOWNLOAD EXCEL FILE",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.warning("Parsed the PDF, but found no loans. Is this a valid Tasdeeq report?")
